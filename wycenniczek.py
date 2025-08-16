import pandas as pd
import os
import glob
import argparse
import re
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from prompt_toolkit import PromptSession
from prompt_toolkit.key_binding import KeyBindings

class CostEstimateManager:
    """Klasa do zarządzania kosztorysem w formacie Excel."""

    def __init__(self, initial_path=None):
        """Inicjalizuje menedżera kosztorysu z pustym DataFrame i flagą modyfikacji."""
        self.filename = None
        self.df = pd.DataFrame(columns=["Pozycja", "Ilość", "Jednostka", "Cena jednostkowa (PLN)", 
                                        "Koszt całkowity (PLN)", "Kategoria", "Opis"])
        self.is_modified = False
        self.current_dir = os.getcwd()

        # Parsowanie ścieżki początkowej
        if initial_path:
            # Normalizacja ścieżki
            initial_path = os.path.normpath(initial_path)
            test_path = initial_path if os.path.isabs(initial_path) else os.path.join(self.current_dir, initial_path)
            # Ochrona przed złośliwymi ścieżkami
            if not os.path.abspath(test_path).startswith(os.path.abspath(self.current_dir)):
                print(f"Ścieżka '{test_path}' wykracza poza bieżący katalog. Użycie bieżącego katalogu.")
                test_path = self.current_dir
            if os.path.isfile(test_path) and test_path.endswith(".xlsx"):
                self.filename = os.path.abspath(test_path)
                self.current_dir = os.path.dirname(test_path) or self.current_dir
                try:
                    os.chdir(self.current_dir)
                    self.df = self.load_cost_estimate()
                    print(f"\n=== Witaj w programie Wycennik! ===")
                    print(f"Bieżący folder: {self.current_dir}")
                    print(f"Kosztorys wczytany z pliku: {os.path.basename(self.filename)}\n")
                    self.display_cost_estimate()
                except Exception as e:
                    print(f"Błąd podczas wczytywania pliku {os.path.basename(self.filename)}: {e}")
                    print("Przechodzenie do trybu interaktywnego.\n")
                    self.filename = None
                    self.df = pd.DataFrame(columns=["Pozycja", "Ilość", "Jednostka", "Cena jednostkowa (PLN)", 
                                                   "Koszt całkowity (PLN)", "Kategoria", "Opis"])
            elif os.path.isdir(test_path):
                try:
                    os.chdir(test_path)
                    self.current_dir = os.getcwd()
                    print(f"\n=== Witaj w programie Wycennik! ===")
                    print(f"Bieżący folder: {self.current_dir}")
                    print("Wybierz plik kosztorysu lub utwórz nowy.")
                except Exception as e:
                    print(f"Błąd podczas zmiany katalogu na {test_path}: {e}")
                    print(f"Przechodzenie do trybu interaktywnego w bieżącym katalogu: {self.current_dir}\n")
            else:
                print(f"Ścieżka '{initial_path}' nie wskazuje na istniejący plik .xlsx ani katalog.")
                print(f"Przechodzenie do trybu interaktywnego w bieżącym katalogu: {self.current_dir}\n")

        # Inicjalizacja PromptSession
        self.prompt_session = PromptSession(multiline=False, enable_history_search=True)
        if not self.filename:
            self.select_initial_file()

    def _get_user_input(self, prompt_message, default="", is_filename=False):
        """Pobiera dane od użytkownika z obsługą strzałek i historii, z sanitizacją."""
        user_input = self.prompt_session.prompt(prompt_message, default=default)
        # Sanitizacja: usuwanie znaków sterujących
        user_input = re.sub(r'[\n\r\t\0]', '', user_input)
        # Ograniczenie długości
        max_length = 255 if is_filename else 1000
        if len(user_input) > max_length:
            print(f"Wprowadzony tekst jest za długi (maks. {max_length} znaków).")
            return ""
        return user_input

    def _get_confirmation(self, prompt_message):
        """Pobiera potwierdzenie (t/n) od użytkownika."""
        return input(prompt_message).lower()

    def _validate_float(self, value, error_message):
        """Waliduje, czy wartość jest liczbą zmiennoprzecinkową w dopuszczalnym zakresie."""
        try:
            val = float(value)
            if val < 0:
                print("Wartość nie może być ujemna.")
                return None
            if val > 1_000_000:
                print("Wartość jest za duża (maks. 1 000 000).")
                return None
            return val
        except ValueError:
            print(error_message)
            return None

    def _validate_filename(self, filename):
        """Waliduje nazwę pliku."""
        if not filename:
            return False
        # Usuwanie znaków sterujących
        filename = re.sub(r'[\n\r\t\0]', '', filename)
        # Sprawdzanie długości
        if len(filename) > 255:
            print("Nazwa pliku jest za długa (maks. 255 znaków).")
            return False
        # Sprawdzanie niedozwolonych znaków
        invalid_chars = r'[<>:"/\\|?*]'
        if re.search(invalid_chars, filename):
            print(f"Nazwa pliku zawiera niedozwolone znaki: {invalid_chars}")
            return False
        # Sprawdzanie sekwencji '..'
        if '..' in filename:
            print("Nazwa pliku nie może zawierać sekwencji '..'.")
            return False
        # Sprawdzanie zarezerwowanych nazw
        reserved_names = ["CON", "PRN", "AUX", "NUL", "COM1", "COM2", "COM3", "COM4",
                         "COM5", "COM6", "COM7", "COM8", "COM9", "LPT1", "LPT2",
                         "LPT3", "LPT4", "LPT5", "LPT6", "LPT7", "LPT8", "LPT9"]
        base_name = filename.split(".")[0].upper()
        if base_name in reserved_names:
            print(f"Nazwa pliku '{base_name}' jest zarezerwowana przez system.")
            return False
        # Dodawanie rozszerzenia .xlsx, jeśli brak
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
        return filename

    def _validate_folder_name(self, folder_name):
        """Waliduje nazwę folderu."""
        if not folder_name:
            return False
        # Usuwanie znaków sterujących
        folder_name = re.sub(r'[\n\r\t\0]', '', folder_name)
        # Sprawdzanie długości
        if len(folder_name) > 255:
            print("Nazwa folderu jest za długa (maks. 255 znaków).")
            return False
        # Sprawdzanie niedozwolonych znaków
        invalid_chars = r'[<>:"/\\|?*]'
        if re.search(invalid_chars, folder_name):
            print(f"Nazwa folderu zawiera niedozwolone znaki: {invalid_chars}")
            return False
        # Sprawdzanie sekwencji '..'
        if '..' in folder_name:
            print("Nazwa folderu nie może zawierać sekwencji '..'.")
            return False
        return folder_name

    def list_excel_files(self):
        """Wyświetla listę plików .xlsx w bieżącym folderze posortowaną według daty modyfikacji."""
        excel_files = [f for f in glob.glob(os.path.join(self.current_dir, "*.xlsx")) if os.path.isfile(f)]
        
        if not excel_files:
            print(f"  Brak plików .xlsx w folderze: {self.current_dir}")
            return []
        
        excel_files = sorted(excel_files, key=lambda f: os.path.getmtime(f), reverse=True)
        
        print(f"\n  Dostępne pliki Excel w folderze {self.current_dir} (posortowane według daty modyfikacji):")
        for idx, file in enumerate(excel_files, 1):
            mod_time = datetime.fromtimestamp(os.path.getmtime(file)).strftime("%Y-%m-%d %H:%M:%S")
            print(f"    {idx}. {os.path.basename(file)} (zmodyfikowany: {mod_time})")
        
        return excel_files

    def list_directories(self):
        """Wyświetla listę folderów w bieżącym katalogu, w tym '..' dla rodzica."""
        dirs = [d for d in glob.glob(os.path.join(self.current_dir, "*")) if os.path.isdir(d)]
        parent_dir = os.path.dirname(self.current_dir)
        dir_list = [".."] + sorted(dirs)
        
        print(f"\n  Foldery w bieżącym katalogu: {self.current_dir}")
        for idx, dir_path in enumerate(dir_list, 1):
            dir_name = os.path.basename(dir_path) if dir_path != ".." else ".."
            print(f"    {idx}. {dir_name}")
        
        return dir_list

    def change_directory(self):
        """Zmienia bieżący katalog na wybrany przez użytkownika."""
        print("\n=== Zmiana folderu ===")
        dir_list = self.list_directories()
        if not dir_list:
            print("  Brak folderów do wyboru.\n")
            return

        while True:
            choice = self._get_user_input("Wpisz numer folderu lub 'q' aby anulować: ")
            if choice.lower() == 'q':
                print("Anulowano. Powrót do menu.\n")
                return
            try:
                dir_idx = int(choice) - 1
                if 0 <= dir_idx < len(dir_list):
                    new_dir = dir_list[dir_idx]
                    if new_dir == "..":
                        new_dir = os.path.dirname(self.current_dir)
                    # Normalizacja i sprawdzanie ścieżki
                    new_dir = os.path.abspath(os.path.normpath(new_dir))
                    if not new_dir.startswith(os.path.abspath(self.current_dir)):
                        print(f"Ścieżka '{new_dir}' wykracza poza bieżący katalog.")
                        continue
                    try:
                        os.chdir(new_dir)
                        self.current_dir = os.getcwd()
                        print(f"Zmieniono folder na: {self.current_dir}\n")
                        self.filename = None
                        self.df = pd.DataFrame(columns=["Pozycja", "Ilość", "Jednostka", "Cena jednostkowa (PLN)", 
                                                       "Koszt całkowity (PLN)", "Kategoria", "Opis"])
                        self.is_modified = False
                        self.list_excel_files()
                        break
                    except Exception as e:
                        print(f"Błąd podczas zmiany folderu: {e}")
                else:
                    print(f"Nieprawidłowy numer. Wybierz od 1 do {len(dir_list)} lub 'q'.")
            except ValueError:
                print("Proszę wpisać poprawną liczbę lub 'q'.")

    def create_directory(self):
        """Tworzy nowy katalog w bieżącym folderze."""
        print("\n=== Tworzenie nowego folderu ===")
        while True:
            folder_name = self._get_user_input("Podaj nazwę nowego folderu ('q' aby anulować): ", is_filename=True)
            if folder_name.lower() == 'q':
                print("Anulowano. Powrót do menu.\n")
                return
            folder_name = self._validate_folder_name(folder_name.strip())
            if not folder_name:
                continue
            folder_path = os.path.abspath(os.path.normpath(os.path.join(self.current_dir, folder_name)))
            if not folder_path.startswith(os.path.abspath(self.current_dir)):
                print(f"Nazwa folderu '{folder_name}' wykracza poza bieżący katalog.")
                continue
            try:
                os.makedirs(folder_path, exist_ok=True)
                print(f"Utworzono folder: {folder_path}\n")
                break
            except OSError as e:
                print(f"Błąd podczas tworzenia folderu '{folder_name}': {e}")
                print("Spróbuj ponownie lub wpisz 'q' aby anulować.")

    def move_cost_estimate(self):
        """Przenosi aktualny kosztorys do wybranego folderu."""
        print("\n=== Przenoszenie kosztorysu ===")
        if not self.filename:
            print("  Brak wczytanego kosztorysu. Najpierw otwórz lub zapisz kosztorys.\n")
            return

        dir_list = self.list_directories()
        if not dir_list:
            print("  Brak folderów do wyboru. Utwórz folder lub zmień bieżący katalog.\n")
            return

        while True:
            choice = self._get_user_input("Wpisz numer folderu docelowego lub 'q' aby anulować: ")
            if choice.lower() == 'q':
                print("Anulowano. Powrót do menu.\n")
                return
            try:
                dir_idx = int(choice) - 1
                if 0 <= dir_idx < len(dir_list):
                    dest_dir = dir_list[dir_idx]
                    if dest_dir == "..":
                        dest_dir = os.path.dirname(self.current_dir)
                    # Normalizacja i sprawdzanie ścieżki
                    dest_dir = os.path.abspath(os.path.normpath(dest_dir))
                    if not dest_dir.startswith(os.path.abspath(self.current_dir)):
                        print(f"Ścieżka '{dest_dir}' wykracza poza bieżący katalog.")
                        continue
                    break
                else:
                    print(f"Nieprawidłowy numer. Wybierz od 1 do {len(dir_list)} lub 'q'.")
            except ValueError:
                print("Proszę wpisać poprawną liczbę lub 'q'.")

        source_path = os.path.abspath(os.path.normpath(self.filename))
        dest_path = os.path.abspath(os.path.normpath(os.path.join(dest_dir, os.path.basename(self.filename))))

        if not dest_path.startswith(os.path.abspath(self.current_dir)):
            print(f"Ścieżka docelowa '{dest_path}' wykracza poza bieżący katalog.")
            return

        if not os.access(dest_dir, os.W_OK):
            print(f"Brak uprawnień do zapisu w folderze docelowym: {dest_dir}")
            print("Powrót do menu.\n")
            return

        if not os.access(source_path, os.R_OK) or not os.access(os.path.dirname(source_path), os.W_OK):
            print(f"Brak uprawnień do odczytu pliku źródłowego lub usunięcia z folderu: {os.path.dirname(source_path)}")
            print("Powrót do menu.\n")
            return

        try:
            if os.path.exists(dest_path):
                print(f"Plik '{os.path.basename(self.filename)}' już istnieje w folderze {dest_dir}.")
                confirm = self._get_confirmation("Czy chcesz nadpisać plik? [t/n]: ")
                if confirm != 't':
                    print("Anulowano. Powrót do menu.\n")
                    return
            shutil.move(source_path, dest_path)
            self.filename = dest_path
            self.current_dir = dest_dir
            os.chdir(self.current_dir)
            print(f"Kosztorys przeniesiony do: {os.path.basename(self.filename)}\n")
        except OSError as e:
            print(f"Błąd podczas przenoszenia pliku: {e}")
            print("Powrót do menu.\n")

    def rename_cost_estimate(self):
        """Zmienia nazwę aktualnego kosztorysu w bieżącym folderze."""
        print("\n=== Zmiana nazwy kosztorysu ===")
        if not self.filename:
            print("  Brak wczytanego kosztorysu. Najpierw otwórz lub zapisz kosztorys.\n")
            return

        default_name = os.path.basename(self.filename)
        while True:
            new_filename = self._get_user_input(
                f"Podaj nową nazwę pliku w folderze {self.current_dir} (Enter dla '{default_name}', 'q' aby anulować): ",
                default=default_name, is_filename=True
            )
            if new_filename.lower() == 'q':
                print("Anulowano. Powrót do menu.\n")
                return
            if not new_filename.strip() or new_filename == default_name:
                print("Nazwa pliku nie zmieniona. Powrót do menu.\n")
                return
            new_filename = self._validate_filename(new_filename.strip())
            if new_filename:
                break
            print("Spróbuj ponownie.")

        new_path = os.path.abspath(os.path.normpath(os.path.join(self.current_dir, new_filename)))
        if not new_path.startswith(os.path.abspath(self.current_dir)):
            print(f"Nowa nazwa pliku '{new_filename}' wykracza poza bieżący katalog.")
            return

        try:
            if os.path.exists(new_path):
                print(f"Plik '{new_filename}' już istnieje w folderze {self.current_dir}.")
                confirm = self._get_confirmation("Czy chcesz nadpisać plik? [t/n]: ")
                if confirm != 't':
                    print("Anulowano. Powrót do menu.\n")
                    return
            os.rename(self.filename, new_path)
            self.filename = new_path
            print(f"Nazwa kosztorysu zmieniona na: {os.path.basename(self.filename)}\n")
        except OSError as e:
            print(f"Błąd podczas zmiany nazwy pliku: {e}")
            print("Powrót do menu.\n")

    def delete_cost_estimate(self):
        """Usuwa wybrany plik kosztorysu z bieżącego folderu."""
        print("\n=== Usuwanie pliku kosztorysu ===")
        excel_files = self.list_excel_files()
        if not excel_files:
            print("  Brak plików do usunięcia.\n")
            return

        while True:
            choice = self._get_user_input("\nWpisz numer pliku do usunięcia lub 'q' aby anulować: ")
            if choice.lower() == 'q':
                print("Anulowano. Powrót do menu.\n")
                return
            try:
                file_idx = int(choice) - 1
                if 0 <= file_idx < len(excel_files):
                    file_to_delete = os.path.abspath(os.path.normpath(excel_files[file_idx]))
                    # Sprawdzanie, czy plik jest w bieżącym katalogu
                    if not file_to_delete.startswith(os.path.abspath(self.current_dir)):
                        print(f"Plik '{os.path.basename(file_to_delete)}' znajduje się poza bieżącym katalogiem.")
                        print("Powrót do menu.\n")
                        return
                    file_name = os.path.basename(file_to_delete)
                    while True:
                        confirm = self._get_confirmation(f"Czy na pewno chcesz usunąć plik '{file_name}'? [t/n]: ")
                        if confirm == 't':
                            try:
                                os.remove(file_to_delete)
                                print(f"Plik '{file_name}' usunięty pomyślnie!\n")
                                if self.filename and os.path.abspath(self.filename) == file_to_delete:
                                    print("Usunięto aktualnie wczytany kosztorys. Tworzenie nowego kosztorysu.\n")
                                    self.filename = None
                                    self.df = pd.DataFrame(columns=["Pozycja", "Ilość", "Jednostka", 
                                                                   "Cena jednostkowa (PLN)", "Koszt całkowity (PLN)", 
                                                                   "Kategoria", "Opis"])
                                    self.is_modified = False
                                break
                            except PermissionError:
                                print(f"Brak uprawnień do usunięcia pliku '{file_name}' lub plik jest w użyciu.")
                                print("Powrót do menu.\n")
                                return
                            except OSError as e:
                                print(f"Błąd podczas usuwania pliku '{file_name}': {e}")
                                print("Powrót do menu.\n")
                                return
                        elif confirm == 'n' or confirm == 'q':
                            print("Anulowano. Powrót do menu.\n")
                            return
                        else:
                            print("Proszę wpisać 't' (tak), 'n' (nie) lub 'q' (anuluj).")
                    break
                else:
                    print(f"Nieprawidłowy numer. Wybierz od 1 do {len(excel_files)} lub 'q'.")
            except ValueError:
                print("Proszę wpisać poprawną liczbę lub 'q'.")

    def select_initial_file(self):
        """Wybiera plik Excel w trybie interaktywnym."""
        excel_files = self.list_excel_files()
        
        if not excel_files:
            print("  Brak plików. Tworzenie nowego kosztorysu.\n")
            self.is_modified = False
            return

        while True:
            choice = self._get_user_input("\nWpisz numer pliku, Enter dla nowego kosztorysu lub 'q' aby anulować: ")
            if choice.lower() == 'q':
                print("Anulowano. Tworzenie nowego kosztorysu.\n")
                self.filename = None
                self.df = pd.DataFrame(columns=["Pozycja", "Ilość", "Jednostka", "Cena jednostkowa (PLN)", 
                                               "Koszt całkowity (PLN)", "Kategoria", "Opis"])
                self.is_modified = False
                return
            if not choice:
                print("Tworzenie nowego kosztorysu.\n")
                self.filename = None
                self.df = pd.DataFrame(columns=["Pozycja", "Ilość", "Jednostka", "Cena jednostkowa (PLN)", 
                                               "Koszt całkowity (PLN)", "Kategoria", "Opis"])
                self.is_modified = False
                return
            
            try:
                file_idx = int(choice) - 1
                if 0 <= file_idx < len(excel_files):
                    self.filename = os.path.abspath(os.path.normpath(excel_files[file_idx]))
                    if not self.filename.startswith(os.path.abspath(self.current_dir)):
                        print(f"Plik '{os.path.basename(self.filename)}' znajduje się poza bieżącym katalogiem.")
                        continue
                    try:
                        self.df = self.load_cost_estimate()
                        print(f"\nKosztorys wczytany z pliku: {os.path.basename(self.filename)}\n")
                        self.is_modified = False
                        self.display_cost_estimate()
                        return
                    except Exception as e:
                        print(f"Błąd podczas wczytywania pliku {os.path.basename(self.filename)}: {e}")
                        print("Tworzenie nowego kosztorysu.\n")
                        self.filename = None
                        self.df = pd.DataFrame(columns=["Pozycja", "Ilość", "Jednostka", "Cena jednostkowa (PLN)", 
                                                       "Koszt całkowity (PLN)", "Kategoria", "Opis"])
                        self.is_modified = False
                        return
                else:
                    print(f"Nieprawidłowy numer. Wybierz od 1 do {len(excel_files)}.")
            except ValueError:
                print("Proszę wpisać poprawną liczbę, Enter lub 'q'.")

    def load_cost_estimate(self):
        """Ładuje kosztorys z pliku lub zgłasza błąd, jeśli plik niepoprawny."""
        if not self.filename or not os.path.exists(self.filename):
            raise Exception(f"Plik {os.path.basename(self.filename)} nie istnieje.")
        
        try:
            df = pd.read_excel(self.filename)
            expected_columns = ["Pozycja", "Ilość", "Jednostka", "Cena jednostkowa (PLN)", 
                               "Koszt całkowity (PLN)", "Kategoria", "Opis"]
            if not all(col in df.columns for col in expected_columns):
                raise Exception(f"Plik {os.path.basename(self.filename)} nie zawiera wszystkich oczekiwanych kolumn.")
            df = df[df["Pozycja"] != "RAZEM"]
            for col in ["Ilość", "Cena jednostkowa (PLN)", "Koszt całkowity (PLN)"]:
                invalid = df[col].isna() | ~pd.to_numeric(df[col], errors='coerce').notna()
                if invalid.any():
                    print(f"Ostrzeżenie: Niepoprawne wartości w kolumnie '{col}' zostały zamienione na 0.")
            df["Ilość"] = pd.to_numeric(df["Ilość"], errors='coerce').fillna(0)
            df["Cena jednostkowa (PLN)"] = pd.to_numeric(df["Cena jednostkowa (PLN)"], errors='coerce').fillna(0)
            df["Koszt całkowity (PLN)"] = pd.to_numeric(df["Koszt całkowity (PLN)"], errors='coerce').fillna(0)
            return df
        except Exception as e:
            raise Exception(f"Błąd podczas wczytywania pliku {os.path.basename(self.filename)}: {e}")

    def open_cost_estimate(self):
        """Wczytuje kosztorys z pliku Excel po numerze."""
        print("\n=== Otwieranie kosztorysu ===")
        excel_files = self.list_excel_files()
        if not excel_files:
            print("  Brak plików do wczytania. Tworzenie nowego kosztorysu.\n")
            self.filename = None
            self.df = pd.DataFrame(columns=["Pozycja", "Ilość", "Jednostka", "Cena jednostkowa (PLN)", 
                                           "Koszt całkowity (PLN)", "Kategoria", "Opis"])
            self.is_modified = False
            return

        while True:
            choice = self._get_user_input("\nWpisz numer pliku lub 'q' aby anulować: ")
            if choice.lower() == 'q':
                print("Anulowano. Powrót do menu.\n")
                return
            try:
                file_idx = int(choice) - 1
                if 0 <= file_idx < len(excel_files):
                    self.filename = os.path.abspath(os.path.normpath(excel_files[file_idx]))
                    if not self.filename.startswith(os.path.abspath(self.current_dir)):
                        print(f"Plik '{os.path.basename(self.filename)}' znajduje się poza bieżącym katalogiem.")
                        continue
                    try:
                        self.df = self.load_cost_estimate()
                        print(f"\nKosztorys wczytany z pliku: {os.path.basename(self.filename)}\n")
                        self.is_modified = False
                        self.display_cost_estimate()
                        break
                    except Exception as e:
                        print(f"Błąd podczas wczytywania pliku {os.path.basename(self.filename)}: {e}")
                        self.filename = None
                        self.df = pd.DataFrame(columns=["Pozycja", "Ilość", "Jednostka", "Cena jednostkowa (PLN)", 
                                                       "Koszt całkowity (PLN)", "Kategoria", "Opis"])
                        self.is_modified = False
                        break
                else:
                    print(f"Nieprawidłowy numer. Wybierz od 1 do {len(excel_files)}.")
            except ValueError:
                print("Proszę wpisać poprawną liczbę lub 'q'.")

    def display_cost_estimate(self):
        """Wyświetla aktualny kosztorys z numerami pozycji."""
        print("\n=== Aktualny kosztorys ===")
        if self.df.empty:
            print("  Kosztorys jest pusty.\n")
        else:
            display_df = self.df.copy()
            display_df.insert(0, "Nr", range(1, len(display_df) + 1))
            print(display_df.to_string(index=False))
            print(f"  Łączny koszt: {self.df['Koszt całkowity (PLN)'].sum():.2f} PLN\n")

    def add_item(self):
        """Dodaje nową pozycję do kosztorysu."""
        print("\n=== Dodawanie nowej pozycji ===")
        while True:
            pozycja = self._get_user_input("Nazwa pozycji ('q' aby anulować): ")
            if pozycja.lower() == 'q':
                print("Anulowano. Powrót do menu.\n")
                return
            if len(pozycja) > 1000:
                print("Nazwa pozycji jest za długa (maks. 1000 znaków).")
                continue
            if pozycja.strip():
                break
            print("Nazwa pozycji nie może być pusta.")

        while True:
            ilosc_input = self._get_user_input("Ilość (Enter dla 1, 'q' aby anulować): ", default="1")
            if ilosc_input.lower() == 'q':
                print("Anulowano. Powrót do menu.\n")
                return
            if not ilosc_input:
                ilosc = 1.0
                break
            ilosc = self._validate_float(ilosc_input, "Proszę podać poprawną wartość liczbową lub 'q'.")
            if ilosc is not None:
                break

        units = ["szt", "m²", "godz", "m³", "kg", "l", "m", "t", "kWh"]
        print("\n  Dostępne jednostki:")
        for idx, unit in enumerate(units, 1):
            print(f"    {idx}. {unit}")
        
        while True:
            unit_choice = self._get_user_input("\nWpisz numer jednostki lub własną jednostkę ('q' aby anulować): ")
            if unit_choice.lower() == 'q':
                print("Anulowano. Powrót do menu.\n")
                return
            if len(unit_choice) > 50:
                print("Jednostka jest za długa (maks. 50 znaków).")
                continue
            try:
                unit_idx = int(unit_choice) - 1
                if 0 <= unit_idx < len(units):
                    jednostka = units[unit_idx]
                    break
                else:
                    print(f"Nieprawidłowy numer. Wybierz od 1 do {len(units)}, wpisz własną jednostkę lub 'q'.")
            except ValueError:
                jednostka = unit_choice
                if jednostka.strip():
                    break
                else:
                    print("Proszę wpisać poprawną jednostkę, numer lub 'q'.")

        while True:
            cena_input = self._get_user_input("Cena jednostkowa (PLN) (Enter dla 0, 'q' aby anulować): ", default="0")
            if cena_input.lower() == 'q':
                print("Anulowano. Powrót do menu.\n")
                return
            if not cena_input:
                cena_jednostkowa = 0.0
                break
            cena_jednostkowa = self._validate_float(cena_input, "Proszę podać poprawną wartość liczbową lub 'q'.")
            if cena_jednostkowa is not None:
                break

        koszt_calkowity = ilosc * cena_jednostkowa

        default_categories = ["Materiały", "Robocizna", "Meble", "Transport"]
        categories = sorted(self.df["Kategoria"].dropna().unique()) if not self.df.empty else default_categories
        if categories:
            print("\n  Dostępne kategorie:")
            for idx, category in enumerate(categories, 1):
                print(f"    {idx}. {category}")
        
        while True:
            cat_choice = self._get_user_input("\nWpisz numer kategorii lub Enter dla własnej ('q' aby anulować): ")
            if cat_choice.lower() == 'q':
                print("Anulowano. Powrót do menu.\n")
                return
            if not cat_choice:
                kategoria = self._get_user_input("Własna kategoria ('q' aby anulować): ")
                if kategoria.lower() == 'q':
                    print("Anulowano. Powrót do menu.\n")
                    return
                if len(kategoria) > 1000:
                    print("Kategoria jest za długa (maks. 1000 znaków).")
                    continue
                if kategoria.strip():
                    break
                else:
                    print("Proszę wpisać poprawną kategorię lub 'q'.")
            try:
                cat_idx = int(cat_choice) - 1
                if 0 <= cat_idx < len(categories):
                    kategoria = categories[cat_idx]
                    break
                else:
                    print(f"Nieprawidłowy numer. Wybierz od 1 do {len(categories)} lub 'q'.")
            except ValueError:
                print("Proszę wpisać poprawną liczbę, Enter lub 'q'.")

        opis = self._get_user_input("Opis (opcjonalny, Enter aby pominąć, 'q' aby anulować): ")
        if opis.lower() == 'q':
            print("Anulowano. Powrót do menu.\n")
            return
        if len(opis) > 1000:
            print("Opis jest za długi (maks. 1000 znaków).")
            opis = opis[:1000]

        new_row = pd.DataFrame({
            "Pozycja": [pozycja],
            "Ilość": [ilosc],
            "Jednostka": [jednostka],
            "Cena jednostkowa (PLN)": [cena_jednostkowa],
            "Koszt całkowity (PLN)": [koszt_calkowity],
            "Kategoria": [kategoria],
            "Opis": [opis]
        })
        self.df = pd.concat([self.df, new_row], ignore_index=True)
        self.is_modified = True
        print("Pozycja dodana pomyślnie!\n")

    def edit_item(self):
        """Edytuje istniejącą pozycję w kosztorysie po numerze pozycji."""
        print("\n=== Edycja pozycji ===")
        if self.df.empty:
            print("  Kosztorys jest pusty. Nie można edytować.\n")
            return

        self.display_cost_estimate()
        while True:
            pozycja_input = self._get_user_input("Wpisz numer pozycji do edycji ('q' aby anulować): ")
            if pozycja_input.lower() == 'q':
                print("Anulowano. Powrót do menu.\n")
                return
            try:
                pozycja_idx = int(pozycja_input) - 1
                if 0 <= pozycja_idx < len(self.df):
                    break
                else:
                    print(f"Nieprawidłowy numer. Wybierz od 1 do {len(self.df)} lub 'q'.")
            except ValueError:
                print("Proszę wpisać poprawną liczbę lub 'q'.")

        print(f"\nEdycja pozycji: {self.df.at[pozycja_idx, 'Pozycja']}")
        pozycja = self.df.at[pozycja_idx, 'Pozycja']
        
        new_pozycja = self._get_user_input(f"Nowa nazwa pozycji (Enter aby pozostawić '{pozycja}', 'q' aby anulować): ", default=pozycja)
        if new_pozycja.lower() == 'q':
            print("Anulowano. Powrót do menu.\n")
            return
        if len(new_pozycja) > 1000:
            print("Nazwa pozycji jest za długa (maks. 1000 znaków).")
            new_pozycja = pozycja
        if not new_pozycja.strip():
            new_pozycja = pozycja

        ilosc_input = self._get_user_input(f"Nowa ilość (Enter aby pozostawić {self.df.at[pozycja_idx, 'Ilość']}, 'q' aby anulować): ", default=str(self.df.at[pozycja_idx, 'Ilość']))
        if ilosc_input.lower() == 'q':
            print("Anulowano. Powrót do menu.\n")
            return
        if not ilosc_input:
            ilosc = self.df.at[pozycja_idx, 'Ilość']
        else:
            ilosc = self._validate_float(ilosc_input, "Nieprawidłowa wartość. Pozostawiono dotychczasową ilość.")
            if ilosc is None:
                ilosc = self.df.at[pozycja_idx, 'Ilość']

        units = ["szt", "m²", "godz", "m³", "kg", "l", "m", "t", "kWh"]
        print("\n  Dostępne jednostki:")
        for idx, unit in enumerate(units, 1):
            print(f"    {idx}. {unit}")
        
        while True:
            unit_choice = self._get_user_input(f"\nWpisz numer jednostki lub własną jednostkę (Enter aby pozostawić '{self.df.at[pozycja_idx, 'Jednostka']}', 'q' aby anulować): ", default=self.df.at[pozycja_idx, 'Jednostka'])
            if unit_choice.lower() == 'q':
                print("Anulowano. Powrót do menu.\n")
                return
            if not unit_choice:
                jednostka = self.df.at[pozycja_idx, 'Jednostka']
                break
            if len(unit_choice) > 50:
                print("Jednostka jest za długa (maks. 50 znaków).")
                continue
            try:
                unit_idx = int(unit_choice) - 1
                if 0 <= unit_idx < len(units):
                    jednostka = units[unit_idx]
                    break
                else:
                    print(f"Nieprawidłowy numer. Wybierz od 1 do {len(units)}, wpisz własną jednostkę lub 'q'.")
            except ValueError:
                jednostka = unit_choice
                if jednostka.strip():
                    break
                else:
                    print("Proszę wpisać poprawną jednostkę, numer lub 'q'.")

        cena_input = self._get_user_input(f"Nowa cena jednostkowa (Enter aby pozostawić {self.df.at[pozycja_idx, 'Cena jednostkowa (PLN)']}, 'q' aby anulować): ", default=str(self.df.at[pozycja_idx, 'Cena jednostkowa (PLN)']))
        if cena_input.lower() == 'q':
            print("Anulowano. Powrót do menu.\n")
            return
        if not cena_input:
            cena_jednostkowa = self.df.at[pozycja_idx, 'Cena jednostkowa (PLN)']
        else:
            cena_jednostkowa = self._validate_float(cena_input, "Nieprawidłowa wartość. Pozostawiono dotychczasową cenę.")
            if cena_jednostkowa is None:
                cena_jednostkowa = self.df.at[pozycja_idx, 'Cena jednostkowa (PLN)']

        koszt_calkowity = ilosc * cena_jednostkowa

        default_categories = ["Materiały", "Robocizna", "Meble", "Transport"]
        categories = sorted(self.df["Kategoria"].dropna().unique()) if not self.df.empty else default_categories
        print("\n  Dostępne kategorie:")
        for idx, category in enumerate(categories, 1):
            print(f"    {idx}. {category}")
        
        while True:
            cat_choice = self._get_user_input(f"\nWpisz numer kategorii lub Enter dla własnej (Enter aby pozostawić '{self.df.at[pozycja_idx, 'Kategoria']}', 'q' aby anulować): ")
            if cat_choice.lower() == 'q':
                print("Anulowano. Powrót do menu.\n")
                return
            if not cat_choice:
                kategoria_input = self._get_user_input(f"Własna kategoria (Enter aby pozostawić '{self.df.at[pozycja_idx, 'Kategoria']}', 'q' aby anulować): ", default=self.df.at[pozycja_idx, 'Kategoria'])
                if kategoria_input.lower() == 'q':
                    print("Anulowano. Powrót do menu.\n")
                    return
                if len(kategoria_input) > 1000:
                    print("Kategoria jest za długa (maks. 1000 znaków).")
                    kategoria = self.df.at[pozycja_idx, 'Kategoria']
                else:
                    kategoria = kategoria_input if kategoria_input.strip() else self.df.at[pozycja_idx, 'Kategoria']
                if kategoria.strip():
                    break
                else:
                    print("Proszę wpisać poprawną kategorię lub 'q'.")
            try:
                cat_idx = int(cat_choice) - 1
                if 0 <= cat_idx < len(categories):
                    kategoria = categories[cat_idx]
                    break
                else:
                    print(f"Nieprawidłowy numer. Wybierz od 1 do {len(categories)} lub 'q'.")
            except ValueError:
                print("Proszę wpisać poprawną liczbę, Enter lub 'q'.")

        opis = self._get_user_input(f"Nowy opis (Enter aby pozostawić '{self.df.at[pozycja_idx, 'Opis']}', 'q' aby anulować): ", default=self.df.at[pozycja_idx, 'Opis'])
        if opis.lower() == 'q':
            print("Anulowano. Powrót do menu.\n")
            return
        if len(opis) > 1000:
            print("Opis jest za długi (maks. 1000 znaków).")
            opis = opis[:1000]

        self.df.at[pozycja_idx, "Pozycja"] = new_pozycja
        self.df.at[pozycja_idx, "Ilość"] = ilosc
        self.df.at[pozycja_idx, "Jednostka"] = jednostka
        self.df.at[pozycja_idx, "Cena jednostkowa (PLN)"] = cena_jednostkowa
        self.df.at[pozycja_idx, "Koszt całkowity (PLN)"] = koszt_calkowity
        self.df.at[pozycja_idx, "Kategoria"] = kategoria
        self.df.at[pozycja_idx, "Opis"] = opis
        self.is_modified = True
        print("Pozycja zaktualizowana pomyślnie!\n")

    def delete_item(self):
        """Usuwa pozycję z kosztorysu po numerze pozycji."""
        print("\n=== Usuwanie pozycji ===")
        if self.df.empty:
            print("  Kosztorys jest pusty. Nie można usunąć.\n")
            return

        self.display_cost_estimate()
        while True:
            pozycja_input = self._get_user_input("Wpisz numer pozycji do usunięcia ('q' aby anulować): ")
            if pozycja_input.lower() == 'q':
                print("Anulowano. Powrót do menu.\n")
                return
            try:
                pozycja_idx = int(pozycja_input) - 1
                if 0 <= pozycja_idx < len(self.df):
                    break
                else:
                    print(f"Nieprawidłowy numer. Wybierz od 1 do {len(self.df)} lub 'q'.")
            except ValueError:
                print("Proszę wpisać poprawną liczbę lub 'q'.")

        pozycja = self.df.at[pozycja_idx, "Pozycja"]
        while True:
            confirm = self._get_confirmation(f"Czy na pewno chcesz usunąć pozycję '{pozycja}'? [t/n]: ")
            if confirm == 't':
                self.df = self.df.drop(index=pozycja_idx).reset_index(drop=True)
                self.is_modified = True
                print(f"Pozycja '{pozycja}' usunięta pomyślnie!\n")
                break
            elif confirm == 'n' or confirm == 'q':
                print("Anulowano. Powrót do menu.\n")
                break
            else:
                print("Proszę wpisać 't' (tak), 'n' (nie) lub 'q' (anuluj).")

    def sort_cost_estimate(self):
        """Sortuje kosztorys według wybranego kryterium."""
        print("\n=== Sortowanie kosztorysu ===")
        if self.df.empty:
            print("  Kosztorys jest pusty. Nie można sortować.\n")
            return

        print("  Opcje sortowania:")
        print("    1. Po nazwie pozycji (A-Z)")
        print("    2. Po koszcie (rosnąco)")
        print("    3. Po kategorii (A-Z)")
        print("    4. Po koszcie (malejąco)")
        choice = self._get_user_input("Wpisz opcję (1-4, 'q' aby anulować): ")
        if choice.lower() == 'q':
            print("Anulowano. Powrót do menu.\n")
            return

        if choice == "1":
            self.df = self.df.sort_values(by="Pozycja")
            print("Kosztorys posortowany po nazwie pozycji.\n")
        elif choice == "2":
            self.df = self.df.sort_values(by="Koszt całkowity (PLN)")
            print("Kosztorys posortowany po koszcie (rosnąco).\n")
        elif choice == "3":
            self.df = self.df.sort_values(by="Kategoria")
            print("Kosztorys posortowany po kategorii.\n")
        elif choice == "4":
            self.df = self.df.sort_values(by="Koszt całkowity (PLN)", ascending=False)
            print("Kosztorys posortowany po koszcie (malejąco).\n")
        else:
            print("Nieprawidłowa opcja.\n")
            return

        self.is_modified = True
        self.display_cost_estimate()

    def filter_cost_estimate(self):
        """Filtruje kosztorys według kategorii lub zakresu kosztów."""
        print("\n=== Filtrowanie kosztorysu ===")
        if self.df.empty:
            print("  Kosztorys jest pusty. Nie można filtrować.\n")
            return

        print("  Opcje filtrowania:")
        print("    1. Po kategorii")
        print("    2. Po zakresie kosztów")
        choice = self._get_user_input("Wpisz opcję (1-2, 'q' aby anulować): ")
        if choice.lower() == 'q':
            print("Anulowano. Powrót do menu.\n")
            return

        if choice == "1":
            categories = sorted(self.df["Kategoria"].dropna().unique())
            if not categories:
                print("  Brak kategorii w kosztorysie.\n")
                return
            print("\n  Dostępne kategorie:")
            for idx, category in enumerate(categories, 1):
                print(f"    {idx}. {category}")
            
            while True:
                cat_choice = self._get_user_input("\nWpisz numer kategorii ('q' aby anulować): ")
                if cat_choice.lower() == 'q':
                    print("Anulowano. Powrót do menu.\n")
                    return
                try:
                    cat_idx = int(cat_choice) - 1
                    if 0 <= cat_idx < len(categories):
                        kategoria = categories[cat_idx]
                        filtered_df = self.df[self.df["Kategoria"] == kategoria]
                        if filtered_df.empty:
                            print(f"  Brak pozycji w kategorii: {kategoria}\n")
                        else:
                            print(f"\n  Pozycje w kategorii {kategoria}:")
                            print(filtered_df.to_string(index=False))
                            print(f"  Łączny koszt w kategorii: {filtered_df['Koszt całkowity (PLN)'].sum():.2f} PLN\n")
                        break
                    else:
                        print(f"Nieprawidłowy numer. Wybierz od 1 do {len(categories)} lub 'q'.")
                except ValueError:
                    print("Proszę wpisać poprawną liczbę lub 'q'.")
        elif choice == "2":
            while True:
                min_koszt_input = self._get_user_input("Minimalny koszt (PLN) ('q' aby anulować): ")
                if min_koszt_input.lower() == 'q':
                    print("Anulowano. Powrót do menu.\n")
                    return
                min_koszt = self._validate_float(min_koszt_input, "Proszę podać poprawną wartość liczbową lub 'q'.")
                if min_koszt is not None:
                    break
            while True:
                max_koszt_input = self._get_user_input("Maksymalny koszt (PLN) ('q' aby anulować): ")
                if max_koszt_input.lower() == 'q':
                    print("Anulowano. Powrót do menu.\n")
                    return
                max_koszt = self._validate_float(max_koszt_input, "Proszę podać poprawną wartość liczbową lub 'q'.")
                if max_koszt is not None:
                    break
            filtered_df = self.df[(self.df["Koszt całkowity (PLN)"] >= min_koszt) & 
                                 (self.df["Koszt całkowity (PLN)"] <= max_koszt)]
            if filtered_df.empty:
                print(f"  Brak pozycji w zakresie kosztów {min_koszt:.2f} - {max_koszt:.2f} PLN\n")
            else:
                print(f"\n  Pozycje w zakresie kosztów {min_koszt:.2f} - {max_koszt:.2f} PLN:")
                print(filtered_df.to_string(index=False))
                print(f"  Łączny koszt w zakresie: {filtered_df['Koszt całkowity (PLN)'].sum():.2f} PLN\n")
        else:
            print("Nieprawidłowa opcja.\n")

    def save_cost_estimate(self):
        """Zapisuje kosztorys do pliku Excel z formatowaniem i kopią zapasową."""
        print("\n=== Zapisywanie kosztorysu ===")
        if self.df.empty:
            print("  Kosztorys jest pusty. Nie można zapisać.\n")
            return

        default_name = os.path.basename(self.filename) if self.filename else "wycennik.xlsx"
        prompt_message = f"Podaj nazwę pliku w folderze {self.current_dir} (Enter dla '{default_name}', 'q' aby anulować): "
        filename_input = self._get_user_input(prompt_message, default=default_name, is_filename=True)
        
        if filename_input.lower() == 'q':
            print("Anulowano. Powrót do menu.\n")
            return
        if not filename_input:
            filename_input = default_name
        
        filename_input = self._validate_filename(filename_input)
        if not filename_input:
            print("Anulowano. Powrót do menu.\n")
            return
        
        self.filename = os.path.abspath(os.path.normpath(os.path.join(self.current_dir, filename_input)))
        if not self.filename.startswith(os.path.abspath(self.current_dir)):
            print(f"Nazwa pliku '{filename_input}' wykracza poza bieżący katalog.")
            return
        
        while True:
            confirm = self._get_confirmation(f"Czy na pewno chcesz zapisać kosztorys do '{filename_input}'? [t/n]: ")
            if confirm == 't':
                break
            elif confirm == 'n' or confirm == 'q':
                print("Anulowano. Powrót do menu.\n")
                return
            else:
                print("Proszę wpisać 't' (tak), 'n' (nie) lub 'q' (anuluj).")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = os.path.join(self.current_dir, f"backup_{timestamp}_{os.path.basename(self.filename)}")
        
        total_cost = self.df["Koszt całkowity (PLN)"].sum()
        summary_row = pd.DataFrame({
            "Pozycja": ["RAZEM"],
            "Ilość": [""],
            "Jednostka": [""],
            "Cena jednostkowa (PLN)": [""],
            "Koszt całkowity (PLN)": [total_cost],
            "Kategoria": [""],
            "Opis": [""]
        })
        df_to_save = pd.concat([self.df, summary_row], ignore_index=True)
        
        try:
            df_to_save.to_excel(self.filename, index=False)
        except PermissionError:
            print(f"Brak uprawnień do zapisu pliku: {os.path.basename(self.filename)}")
            return
        except OSError as e:
            if e.errno == 28:  # errno.ENOSPC - brak miejsca na urządzeniu
                print("Brak miejsca na dysku. Nie można zapisać pliku.")
            else:
                print(f"Błąd podczas zapisu pliku: {e}")
            return
        
        wb = load_workbook(self.filename)
        ws = wb.active
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        total_font = Font(bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        for col_idx, column in enumerate(df_to_save.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                if col_idx in [2, 3, 4, 5]:
                    cell.alignment = center_align
                    if col_idx in [2, 4, 5]:
                        cell.number_format = '#,##0.00'
                else:
                    cell.alignment = left_align
                if row_idx == ws.max_row:
                    cell.font = total_font
        
        for col_idx, column in enumerate(df_to_save.columns, 1):
            max_length = max(len(str(column)), 10)
            for value in df_to_save[column]:
                try:
                    max_length = max(max_length, len(str(value)))
                except:
                    pass
                adjusted_width = max_length * 1.2
                ws.column_dimensions[get_column_letter(col_idx)].width = max(adjusted_width, 10)
        
        wb.save(self.filename)
        
        if os.path.exists(self.filename) and os.path.getsize(self.filename) > 0:
            try:
                df_to_save.to_excel(backup_filename, index=False)
                wb_backup = load_workbook(backup_filename)
                ws_backup = wb_backup.active
                
                for col_idx, column in enumerate(df_to_save.columns, 1):
                    cell = ws_backup.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                for row_idx in range(2, ws_backup.max_row + 1):
                    for col_idx in range(1, ws_backup.max_column + 1):
                        cell = ws_backup.cell(row=row_idx, column=col_idx)
                        cell.border = border
                        if col_idx in [2, 3, 4, 5]:
                            cell.alignment = center_align
                            if col_idx in [2, 4, 5]:
                                cell.number_format = '#,##0.00'
                        else:
                            cell.alignment = left_align
                        if row_idx == ws_backup.max_row:
                            cell.font = total_font
                for col_idx, column in enumerate(df_to_save.columns, 1):
                    max_length = max(len(str(column)), 10)
                    for value in df_to_save[column]:
                        try:
                            max_length = max(max_length, len(str(value)))
                        except:
                            pass
                    adjusted_width = max_length * 1.2
                    ws_backup.column_dimensions[get_column_letter(col_idx)].width = max(adjusted_width, 10)
                wb_backup.save(backup_filename)
                print(f"Utworzono kopię zapasową: {os.path.basename(backup_filename)}")
            except Exception as e:
                print(f"Błąd podczas tworzenia kopii zapasowej: {e}")
        
        self.is_modified = False
        print(f"Kosztorys zapisany do: {os.path.basename(self.filename)}\n")

    def run(self):
        """Główna pętla programu z menu głównym."""
        while True:
            print("\n=== Wycennik - Zarządzanie kosztorysem ===")
            print(f"  Bieżący folder: {self.current_dir}")
            print("  1. Otwórz kosztorys z pliku")
            print("  2. Wyświetl kosztorys")
            print("  3. Dodaj pozycję")
            print("  4. Edytuj pozycję")
            print("  5. Usuń pozycję")
            print("  6. Sortuj kosztorys")
            print("  7. Filtruj kosztorys")
            print("  8. Zapisz kosztorys")
            print("  9. Zmień folder")
            print("  10. Utwórz nowy folder")
            print("  11. Przenieś kosztorys do folderu")
            print("  12. Zmień nazwę kosztorysu")
            print("  13. Usuń plik kosztorysu")
            print("  14. Wyjdź")
            choice = self._get_user_input("\nWpisz opcję (1-14): ")
            print()

            if choice == "1":
                self.open_cost_estimate()
            elif choice == "2":
                self.display_cost_estimate()
            elif choice == "3":
                self.add_item()
            elif choice == "4":
                self.edit_item()
            elif choice == "5":
                self.delete_item()
            elif choice == "6":
                self.sort_cost_estimate()
            elif choice == "7":
                self.filter_cost_estimate()
            elif choice == "8":
                self.save_cost_estimate()
            elif choice == "9":
                self.change_directory()
            elif choice == "10":
                self.create_directory()
            elif choice == "11":
                self.move_cost_estimate()
            elif choice == "12":
                self.rename_cost_estimate()
            elif choice == "13":
                self.delete_cost_estimate()
            elif choice == "14":
                if self.is_modified:
                    while True:
                        confirm = self._get_confirmation("Czy na pewno chcesz wyjść bez zapisywania zmian? [t/n]: ")
                        if confirm == 't':
                            print("Zakończenie programu.\n")
                            return
                        elif confirm == 'n' or confirm == 'q':
                            print("Powrót do menu.\n")
                            break
                        else:
                            print("Proszę wpisać 't' (tak), 'n' (nie) lub 'q' (anuluj).")
                else:
                    print("Zakończenie programu.\n")
                    return
            else:
                print("Nieprawidłowa opcja. Wybierz od 1 do 14.\n")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Wycennik - Zarządzanie kosztorysem")
    parser.add_argument("path", type=str, nargs='?', default=None, help="Ścieżka do pliku .xlsx lub katalogu")
    args = parser.parse_args()
    manager = CostEstimateManager(initial_path=args.path)
    manager.run()
